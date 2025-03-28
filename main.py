import curses
from tradingview_ta import TA_Handler, Interval
import time
import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime

def calcular_variacao_percentual(preco_antigo, preco_novo):
    # Calcula la variación porcentual entre dos precios
    if preco_antigo == 0:
        return 0
    return ((preco_novo - preco_antigo) / preco_antigo) * 100

# Definir los tickers de las criptomonedas a monitorar
cryptos = ['BTCUSD', 'SOLUSD', 'XRPUSD', 'BNBUSD', 'ETHUSD', 'XLMUSD', 'TRXUSD', 'DOTUSD']

# Definir los tickers de las criptomonedas a excluir de BoostUP y BoostDOWN
# Ejemplo: excepciones = ['BTCUSD', 'ETHUSD']

excepciones = ['LUNAUSD']

def obtener_todos_los_tickers():
    # Esta función debería devolver una lista de todos los tickers de criptoactivos disponibles.
    # Aquí se usa un ejemplo estático, pero en un caso real, se debería obtener dinámicamente.
    return ['BTCUSD', 'SOLUSD', 'XRPUSD', 'BNBUSD', 'ETHUSD', 'XLMUSD', 'TRXUSD', 'DOTUSD']

def calcular_boost(precos_anteriores, precos_novos):
    todos_los_tickers = obtener_todos_los_tickers()
    maior_subida = {"ticker": None, "variacao": -float('inf'), "preco": 0}
    maior_baixa = {"ticker": None, "variacao": float('inf'), "preco": 0}
    
    for crypto in todos_los_tickers:
        if crypto in excepciones:
            continue  # Saltar los tickers en la lista de excepciones
        
        preco_antigo, _ = precos_anteriores.get(crypto, (0, 0))
        preco_novo, _ = precos_novos.get(crypto, (0, 0))
        variacao_preco = calcular_variacao_percentual(preco_antigo, preco_novo)
        
        if variacao_preco > maior_subida["variacao"]:
            maior_subida = {"ticker": crypto, "variacao": variacao_preco, "preco": preco_novo}
        
        if variacao_preco < maior_baixa["variacao"]:
            maior_baixa = {"ticker": crypto, "variacao": variacao_preco, "preco": preco_novo}
    
    return maior_subida, maior_baixa

def obter_dados(crypto):
    # Obtiene los datos de cierre y volumen de una criptomoneda
    handler = TA_Handler(
        symbol=crypto,
        exchange="BINANCE",
        screener="crypto",
        interval=Interval.INTERVAL_1_MINUTE   # Opciones posibles: INTERVAL_1_MINUTE, INTERVAL_5_MINUTES, INTERVAL_15_MINUTES, INTERVAL_30_MINUTES, INTERVAL_1_HOUR, INTERVAL_4_HOURS, INTERVAL_1_DAY, INTERVAL_1_WEEK, INTERVAL_1_MONTH
    )
    analysis = handler.get_analysis().indicators
    return analysis['close'], analysis['volume']

def inicializar_datos_historicos(filepath):
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos Historicos"
        ws.append(["Fecha/Hora", "BoostUP", "BoostDOWN"] + cryptos)
        wb.save(filepath)
    else:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        existing_tickers = [ws.cell(row=1, column=i).value for i in range(2, ws.max_column + 1)]
        for crypto in cryptos:
            if crypto not in existing_tickers:
                ws.cell(row=1, column=ws.max_column + 1, value=crypto)
        wb.save(filepath)

def actualizar_datos_historicos(filepath, datos, boost_up, boost_down):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    new_row = [current_time, f"{boost_up['ticker']} {boost_up['preco']:.2f} ({boost_up['variacao']:+.2f}%)", f"{boost_down['ticker']} {boost_down['preco']:.2f} ({boost_down['variacao']:+.2f}%)"] + [datos.get(crypto, (None,))[0] for crypto in cryptos]
    ws.append(new_row)
    wb.save(filepath)

def main(stdscr):
    try:
        curses.start_color()
        curses.init_pair(1, curses.COLOR_WHITE, curses.COLOR_BLUE)
        curses.init_pair(2, curses.COLOR_WHITE, curses.COLOR_RED)
        curses.init_pair(3, curses.COLOR_WHITE, curses.COLOR_GREEN)

        # Inicializar archivo de datos históricos
        datos_historicos_filepath = "DatosHistoricos.xlsx"
        inicializar_datos_historicos(datos_historicos_filepath)

        # Obtener precios iniciales
        precos_iniciais = {crypto: obter_dados(crypto) for crypto in cryptos}
        precos_anteriores = precos_iniciais.copy()
        curses.curs_set(0)  # Ocultar el cursor
        stdscr.nodelay(1)  # No bloquear en getch()
        stdscr.timeout(60000)  # Actualizar cada 60000 ms (1 MINUTO = 60000 ms)

        linha = 0
        while True:
            # Obtener nuevos precios
            precos_novos = {crypto: obter_dados(crypto) for crypto in cryptos}
            boost_up, boost_down = calcular_boost(precos_anteriores, precos_novos)
            stdscr.clear()
            for i, crypto in enumerate(cryptos):
                preco_inicial, volume_inicial = precos_iniciais[crypto]
                preco_novo, volume_novo = precos_novos[crypto]
                preco_antigo, volume_antigo = precos_anteriores[crypto]
                variacao_preco = calcular_variacao_percentual(preco_antigo, preco_novo)
                variacao_volume = calcular_variacao_percentual(volume_antigo, volume_novo)

                try:
                    stdscr.addstr(linha, 0, f'{crypto}:', curses.color_pair(1))
                    stdscr.addstr(linha + 1, 0, f'  Precio Inicial: {preco_inicial:.2f}', curses.color_pair(2))
                    stdscr.addstr(linha + 2, 0, f'  Precio Actual: {preco_novo:.2f}', curses.color_pair(2))
                    stdscr.addstr(linha + 3, 0, f'  Variación Precio: {variacao_preco:.2f}%', curses.color_pair(3))
                    stdscr.addstr(linha + 4, 0, f'  Variación Volumen: {variacao_volume:.2f}%', curses.color_pair(3))
                except curses.error:
                    pass  # Ignorar errores causados por escribir fuera de la ventana

                linha += 6

            # Mostrar BoostUP y BoostDOWN
            try:
                stdscr.addstr(linha, 0, f'BoostUP: {boost_up["ticker"]} {boost_up["preco"]:.2f} ({boost_up["variacao"]:+.2f}%)', curses.color_pair(3))
                stdscr.addstr(linha + 1, 0, f'BoostDOWN: {boost_down["ticker"]} {boost_down["preco"]:.2f} ({boost_down["variacao"]:+.2f}%)', curses.color_pair(2))
            except curses.error:
                pass  # Ignorar errores causados por escribir fuera de la ventana

            stdscr.refresh()
            precos_anteriores = precos_novos
            linha = 0

            # Actualizar datos históricos
            datos_historicos = {crypto: (precos_novos[crypto][0], calcular_variacao_percentual(precos_anteriores[crypto][0], precos_novos[crypto][0])) for crypto in cryptos}
            actualizar_datos_historicos(datos_historicos_filepath, datos_historicos, boost_up, boost_down)

            if stdscr.getch() != -1:
                break
    except KeyboardInterrupt:
        pass  # Permitir salir con Ctrl+C sin errores

curses.wrapper(main)